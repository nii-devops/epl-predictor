#from turtle import title
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, abort, send_file, make_response, current_app
from flask_login import login_user, login_required, logout_user, current_user
from datetime import date, datetime
from werkzeug.security import generate_password_hash, check_password_hash
import string
from sqlalchemy.engine import url
#from . import TWILIO_PHONE_NUMBER, twilio_client, mail, s
from itsdangerous import SignatureExpired, BadSignature, BadTimeSignature

from .models import *
from .forms import *
from wtforms import FieldList, FormField
from . import oauth, ADMINS, mail, s, TOKEN_SALT
import os
import io
from pprint import pprint
import pandas as pd
from fpdf import FPDF
import traceback
from sqlalchemy import inspect as sql_inspect
import inspect
import openpyxl
import secrets
from flask_mail import Message
from twilio.rest import Client



bp = Blueprint('main', __name__)



ADMIN_EMAILS = os.getenv('ADMIN_EMAILS')

MODELS = [User, MatchWeek, Season, Team, MatchWeekPoint, Fixture, Prediction]

weeks = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 
        21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38]



#############################
# Twilio Configuration
TWILIO_ACCOUNT_SID = os.environ.get('TWILIO_ACCOUNT_SID')
TWILIO_AUTH_TOKEN = os.environ.get('TWILIO_AUTH_TOKEN')
TWILIO_PHONE_NUMBER = os.environ.get('TWILIO_PHONE_NUMBER')
twilio_client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)



##########################################################
# Define the inactivity timeout in seconds (5 minutes = 300 seconds)
INACTIVITY_TIMEOUT_SECONDS = 300

bp.before_request
def make_session_permanent_and_check_inactivity():
    session.permanent = True

    if 'username' in session:
        # Get the time of the last activity from the session.
        last_activity = session.get('last_activity', None)

        if last_activity is None:
            session['last_activity'] = datetime.datetime.now()
            return

        time_since_last_activity = datetime.datetime.now() - last_activity
        
        if time_since_last_activity.total_seconds() > INACTIVITY_TIMEOUT_SECONDS:
            # Clear the entire session to effectively log the user out.
            session.clear()
            flash('You have been logged out due to inactivity.', 'warning')
            return redirect(url_for('main.login'))
        session['last_activity'] = datetime.datetime.now() 
    return


def _log_exception(e):
    try:
        current_app.logger.exception(str(e))
    except Exception:
        print("Exception logging failed:")
        traceback.print_exc()


@bp.route('/')
def index():
    try:
        now = datetime.utcnow()
        print(now)
        active_match_weeks = MatchWeek.query.filter(MatchWeek.predictions_close_time > now).all()
        return render_template('index.html', active_match_weeks=active_match_weeks)
    except Exception as e:
        _log_exception(e)
        flash('An error occurred while loading the homepage.', 'error')
        # safe fallback: empty list
        return render_template('index.html', active_match_weeks=[])



############################################
# TWILIO COMMANDS
def generate_random_code(length=6):
    """Generate a random alphanumeric code."""
    characters = string.ascii_uppercase + string.digits
    return ''.join(secrets.choice(characters) for _ in range(length))


def send_sms(to_number, message):
    """Sends an SMS message using Twilio."""
    try:
        twilio_client.messages.create(
            to=to_number,
            from_=TWILIO_PHONE_NUMBER,
            body=message
        )
        print(f"SMS sent to {to_number}")
    except Exception as e:
        print(f"Error sending SMS: {e}")



# New API endpoint to serve user data
@bp.route('/api/users', methods=['GET'])
def get_users_api():
    users = User.query.order_by(User.name).all()
    user_data = [{
        'id': user.id,
        'name': user.name,
        'nickname': user.nickname
    } for user in users]
    return jsonify(user_data)



@bp.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        email = form.email.data
        password = form.password.data
        user = User.query.filter_by(email=email).first()
        if user:
            if check_password_hash(user.password, password):
                login_user(user)
                return redirect(url_for('main.index'))
        else:
            flash('User not found! Register.', 'info')
            return redirect(url_for('main.register'))
    return render_template('admin/login.html', heading='Login', title='Login', form=form)




@bp.route('/register', methods=['GET', 'POST'])
def register():
    form = RegisterForm()
    if form.validate_on_submit():
        name = form.name.data
        email = form.email.data
        phone_no = form.phone_no.data
        nickname = form.nickname.data
        password = form.password_1.data
        

        # Write data to database...
        if User.query.filter_by(email=email, phone_no=phone_no).first():
            flash('User exists', 'info')
            return redirect(url_for('main.login'))
        
        if email in ADMINS:
            try:
                db.session.add(
                    User(
                        name=name, email=email, phone_no=phone_no,
                        nickname=nickname, is_admin=True,
                        password=generate_password_hash(password, method='scrypt', salt_length=8)
                    )
                )
                db.session.commit()
                flash('User created', 'success')
                return redirect(url_for('main.login'))
            except Exception as e:
                _log_exception(e)
                flash('An error occurred loading the login page.', 'error')
                return redirect(url_for('main.index'))
            
        else:
            try:
                db.session.add(
                    User(
                        name=name, email=email, phone_no=phone_no,
                        nickname=nickname, is_admin=False,
                        password=generate_password_hash(password, method='scrypt', salt_length=8)
                    )
                )
                db.session.commit()
                flash('User created', 'success')
                return redirect(url_for('main.login'))
            except Exception as e:
                _log_exception(e)
                flash('An error occurred loading the login page.', 'error')
                return redirect(url_for('main.index'))
    return render_template('admin/register.html', heading='Register', title='Register', form=form)





# The following routes require OAuth and app context, which will be handled in __init__.py
# Placeholders for now:
@bp.route('/authorize/google')
def google_auth():
    try:
        redirect_uri = url_for('main.google_callback', _external=True)
        #redirect_uri = os.getenv('REDIRECT_URI')  # Use environment variable if set
        return oauth.google.authorize_redirect(redirect_uri)
    except Exception as e:
        _log_exception(e)
        flash('Error starting Google authentication.', 'error')
        return redirect(url_for('main.login'))



@bp.route('/authorize/google/callback')
def google_callback():
    try:
        token = oauth.google.authorize_access_token()
        user_info = token.get('userinfo')

        if not user_info:
            flash('Authentication failed', 'error')
            return redirect(url_for('main.login'))

        # Check if user exists by Google ID
        user = User.query.filter_by(google_id=user_info['sub']).first()

        # If not found by google_id, check by email
        if not user:
            user = User.query.filter_by(email=user_info['email']).first()

        if not user:
            # First-time Google login → redirect to set_nickname
            is_admin = 1 if user_info['email'] in ADMIN_EMAILS else 0
            return redirect(url_for(
                'main.set_nickname',
                google_id=user_info['sub'],
                email=user_info['email'],
                name=user_info['name'],
                is_admin=is_admin
            ))

        # If found, update details and log them in
        user.google_id = user_info['sub']
        user.name = user_info['name']
        db.session.commit()

        login_user(user)
        flash('Successfully logged in!', 'success')
        return redirect(url_for('main.index'))
    except Exception as e:
        _log_exception(e)
        flash('Google authentication callback failed.', 'error')
        return redirect(url_for('main.login'))


@bp.route('/user/set-nickname', methods=['GET', 'POST'])
def set_nickname():
    try:
        google_id = request.args.get('google_id')
        email = request.args.get('email')
        name = request.args.get('name')
        is_admin = int(request.args.get('is_admin', 0))

        form = NameForm()

        # Pre-fill form on GET
        if request.method == 'GET':
            form.name.data = name
            form.email.data = email
            form.google_id.data = google_id

        if form.validate_on_submit():
            # Save user to DB
            user = User(
                nickname=form.nickname.data,
                email=form.email.data,
                name=form.name.data,
                google_id=google_id,
                is_admin=True if is_admin == 1 else False
            )
            db.session.add(user)
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                _log_exception(e)
                flash('Error saving account. Please try again.', 'error')
                return render_template('admin/form.html', title='Set Nickname', heading='Set Nickname', form=form)

            login_user(user)
            flash('Account created and logged in!', 'success')
            return redirect(url_for('main.index'))

        return render_template('admin/form.html', title='Set Nickname', heading='Set Nickname', form=form)
    except Exception as e:
        _log_exception(e)
        flash('An error occurred setting nickname.', 'error')
        return redirect(url_for('main.login'))



@bp.route('/logout')
@login_required
def logout():
    try:
        logout_user()
        flash('You have been logged out', 'info')
        return redirect(url_for('main.index'))
    except Exception as e:
        _log_exception(e)
        flash('Error during logout.', 'error')
        return redirect(url_for('main.index'))



########################################
##### ADMIN DASHBOARD ###############

@bp.route('/admin')
@login_required
def admin_dashboard():
    now = datetime.utcnow()
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('main.index'))
    
    users = User.query.order_by(User.id).all()
    match_weeks = MatchWeek.query.order_by(MatchWeek.id).all()
    for mw in match_weeks:
        print(mw.predictions_close_time)
    seasons = Season.query.order_by(Season.id).all()
    teams = Team.query.order_by(Team.id).all()
    fixtures = Fixture.query.order_by(Fixture.id).all()

    return render_template('admin/dashboard.html', title='Admin Panel', now=now,
        users=users, match_weeks=match_weeks, seasons=seasons, teams=teams, fixtures=fixtures)



@bp.route('/user/toggle-admin/<int:user_id>', methods=['POST', 'GET'])
@login_required
def toggle_admin(user_id):
    if not current_user.is_admin:
        flash("Access denied. Only admins can modify admin roles.", "danger")
        return redirect(url_for('main.index'))

    user = User.query.get_or_404(user_id)

    # Prevent last admin from being demoted
    if user.is_admin and User.query.filter_by(is_admin=True).count() == 1:
        flash("You cannot remove the last admin!", "warning")
        return redirect(request.referrer or url_for('main.admin_dashboard'))

    user.is_admin = not user.is_admin

    try:
        db.session.commit()
        flash(f"{user.name} admin status updated to {user.is_admin}.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error toggling admin: {e}", "danger")

    return redirect(request.referrer or url_for('main.index'))



########################################
##### PASSWORD RESET ###############

@bp.route('/request_password_reset', methods=['GET', 'POST'])
def request_password_reset():
    form = ResetEmailForm()
    if form.validate_on_submit():
        email = form.email.data
        user = User.query.filter_by(email=email).first()

        if user:
            # Generate a time-sensitive token with the user's email
            #token = s.dumps(email, salt='password-reset-salt')
            token = s.dumps(email, salt=TOKEN_SALT)
            
            #token = current_app.token_serializer.dumps(email, salt='password-reset-salt')
            
            # Create the reset link
            reset_link = url_for('main.reset_password', token=token, _external=True)

            # Send the email
            msg = Message(
                "Password Reset Request",
                sender=current_app.config['MAIL_USERNAME'],
                recipients=[email]
            )
            msg.body = f"Hello {user.name}, \nPlease click the following link to reset your password: \n{reset_link}"
            
            try:
                mail.send(msg)
                flash('A password reset link has been sent to your email.', 'success')
                return redirect(url_for('main.index'))
            except Exception as e:
                flash(f'Failed to send email: {e}', 'danger')
                return redirect(url_for('main.index'))
        else:
            flash("Email address not found.", 'info')
            return redirect(url_for('main.index'))
    
    return render_template('admin/form.html', form=form, heading='Request Password Reset', title='Request Password Reset')


@bp.route('/reset_password/<token>', methods=['GET'])
def reset_password(token):
    try:
        # Load the token with an expiration time (e.g., 3600 seconds = 1 hour)

        email = s.loads(token, salt=TOKEN_SALT, max_age=3600)
        
        # If the token is valid and not expired, redirect to the password change form
        return redirect(url_for('main.show_password_reset_form', token=token))
        
    except SignatureExpired:
        flash("The password reset link has expired. Please request a new one.", 'danger')
        return redirect(url_for('main.request_password_reset'))
    except (BadSignature, BadTimeSignature):
        # This will catch any cryptographic signature mismatch or timestamp issues
        flash("Invalid token signature. Please request a new link.", 'danger')
        return redirect(url_for('main.request_password_reset'))
    except Exception as e:
        # A generic fallback for any other unexpected error
        flash(f"An unexpected error: {e}.", 'danger')
        return redirect(url_for('main.request_password_reset'))


@bp.route('/reset_password_form/<token>', methods=['GET', 'POST'])
def show_password_reset_form(token):
    form = PasswordResetForm()
    if form.validate_on_submit():
        # This part handles the form submission from the user
        new_password = form.password_1.data
        
        try:
            # Verify the token again before changing the password
            email = s.loads(token, salt=TOKEN_SALT, max_age=3600)

            user = User.query.filter_by(email=email).first()

            if user:
                # Hash the new password before saving it
                user.password = generate_password_hash(new_password, method='scrypt', salt_length=8)

                # Step 4: Commit the change to the database
                db.session.commit()
                flash("Your password has been reset successfully!", 'success')
                return redirect(url_for('main.index'))

        except SignatureExpired:
            flash("The password reset link has expired. Please request a new one.", 'danger')
            return redirect(url_for('main.index'))
            # return "The password reset link has expired. Please request a new one.", 400
        except Exception:
            flash("Invalid token.", 'danger')
            return redirect(url_for('main.index'))
            # return "Invalid token.", 400
            
    return render_template('admin/form.html', form=form, token=token, heading='Create New Password', title='Create New Password')



"""
@bp.route('/password_reset/verify', methods=['POST'])
def verify_code():
    data = request.get_json()
    token = data.get('token')
    reset_code = data.get('code')
    new_password = data.get('new_password')

    if not token or not reset_code or not new_password:
        return jsonify({"message": "Missing token, code or new password"}), 400

    # Step 4: Verification
    try:
        # Check if the token is valid and not expired
        identifier = s.loads(token, max_age=3600)  # max_age in seconds
        
        # Check if the code and token match in the database
        stored_token_data = tokens_db.get(token)
        if not stored_token_data or stored_token_data['code'] != reset_code:
            return jsonify({"message": "Invalid code or token"}), 400
        
        # Check if the code has expired
        if datetime.datetime.now() > stored_token_data['expiry']:
            del tokens_db[token] # Invalidate the token
            return jsonify({"message": "Reset code has expired"}), 400

        # Step 5: Reset password
        # In a real app, update the user's password in the database
        users_db[identifier]['password'] = new_password
        
        # Invalidate the token after successful use
        del tokens_db[token]
        
        return jsonify({"message": "Password has been successfully reset."}), 200

    except SignatureExpired:
        return jsonify({"message": "The reset token has expired"}), 400
    except Exception as e:
        return jsonify({"message": "Invalid token"}), 400
"""

@bp.route('/user/password/self-reset', methods=['GET', 'POST'])
#@login_required
def password_self_reset():
    user = User.query.get_or_404()
    return redirect(url_for('main.admin_dashboard'))



@bp.route('/user/reset-password/<int:user_id>', methods=['GET', 'POST'])
@login_required
def reset_user_password(user_id):
    user = User.query.get_or_404(user_id)
    return redirect(url_for('main.admin_dashboard'))



@bp.route('/user/telephone-number', methods=['GET', 'POST'])
#@login_required
def telephone_number():
    form = TelephoneNumberForm()
    if form.validate_on_submit():
        country_code = form.country_code.data
        phone_number = form.phone_number.data
        full_number = f'{country_code}{phone_number}'
        print(full_number)
        return redirect(request.referrer)
    return render_template('admin/telephone_form.html', title='Telephone Number', heading='Telephone Number', form=form)



@bp.route('/users/telephone-number/standardize', methods=['GET', 'POST'])
#@login_required
def standardize_telephone_numbers():
    users = User.query.all()
    gh_code = '+233'
    if not users:
        flash('There are no users', 'info')
        return redirect(request.referrer)
    for user in users:
        user_num = user.phone_no
        if user_num[0] == '0':
            user_num = user_num[1::]
            phone_num = f'{gh_code}{user_num}'
        print(phone_num)
    return redirect(request.referrer)
    #return render_template('admin/telephone_form.html', title='Telephone Number', heading='Telephone Number')




@bp.route('/user/edit-user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    if not current_user.is_admin:
        flash('Access denied!', 'info')
        return redirect(url_for('main.index'))

    user = User.query.get_or_404(user_id)
    if not user:
        flash('User not found!', 'danger')
        return redirect(request.referrer)

    form = EditUserForm()

    if request.method == "GET":
        form.name.data = user.name
        form.email.data = user.email
        form.phone_no.data = user.phone_no
        form.nickname.data = user.nickname

    if form.validate_on_submit():
        try:
            user.name = form.name.data
            user.email = form.email.data
            user.phone_no = form.phone_no.data
            user.nickname = form.nickname.data
            db.session.commit()
            flash('User updated successfully.', 'success')

        except Exception as e:
            _log_exception
            db.session.rollback()
            flash(f'Error occurred: {e}')
        return redirect(url_for('main.admin_dashboard'))

    return render_template('admin/form.html', form=form, title='Edit User', heading='Edit User Details')


@bp.route('/user/delete-user/<int:user_id>')
@login_required
def delete_user(user_id):
    if not current_user.is_admin:
        flash('Access denied!', 'info')
        return redirect(url_for('main.index'))

    user = User.query.get_or_404(user_id)
    if not user:
        flash('User not found!', 'danger')
        return redirect(request.referrer)

    # Query all other users except this one
    other_users = User.query.filter(User.id != user.id).all()

    # Prevent deleting the last admin
    if user.is_admin and not any(u.is_admin for u in other_users):
        flash('You cannot delete the last admin. Assign another admin first.', 'danger')
        return redirect(request.referrer)

    try:
        name = user.name
        db.session.delete(user)
        db.session.commit()
        flash(f"{name}'s account deleted.", 'success')

    except Exception as e:
            _log_exception
            db.session.rollback()
            flash(f'Error occurred: {e}')
    return redirect(request.referrer or url_for('main.index'))



@bp.route('/create_weeks', methods=['GET', 'POST'])
@login_required
def create_weeks():
    try:
        if not current_user.is_admin:
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('main.index'))
        
        # Check if weeks already exist
        for week_num in range(1, 39):
            try:
                if Week.query.filter_by(week_number=week_num).first():
                    flash(f'Week {week_num} already exists. Skipping creation.', 'warning')
                    continue    
                
                week = Week(week_number=week_num)   
                db.session.add(week)
            except Exception as e:
                _log_exception(e)
                flash(f'Error creating week {week_num}: {e}', 'error')
        try:
            db.session.commit()
            flash('Weeks created successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            _log_exception(e)
            flash('Error committing weeks to the database.', 'error')
        return redirect(url_for('main.admin_dashboard'))
    except Exception as e:
        _log_exception(e)
        flash('An unexpected error occurred creating weeks.', 'error')
        return redirect(url_for('main.admin_dashboard'))



@bp.route('/admin/teams/populate', methods=['GET', 'POST'])
@login_required
def populate_teams():
    try:
        if not current_user.is_admin:
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('main.index'))  
        for team in EPL_TEAMS:
            for short_name, name in team.items():
                try:
                    if not Team.query.filter_by(name=name).first():
                        new_team = Team(name=name, short_name=short_name)
                        db.session.add(new_team)
                    else:
                        current_app.logger.info(f"Team {name} already exists, skipping.")
                        return redirect(request.referrer)
                except Exception as e:
                    _log_exception(e)
                    flash(f"Error checking/creating team {name}: {e}", 'error')
        try:
            db.session.commit()
            flash('Teams populated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            _log_exception(e)
            flash('Error committing teams to DB.', 'error')
        return redirect(request.referrer)
    except Exception as e:
        _log_exception(e)
        flash('Error populating teams.', 'error')
        return redirect(url_for('main.admin_dashboard'))



@bp.route('/admin/team/edit/<int:team_id>', methods=['GET', 'POST'])
@login_required
def edit_team(team_id):
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(request.referrer)
    try:
        now = datetime.utcnow()
        team = Team.query.get_or_404(team_id)

        form = CreateTeamForm()
        if request.method == 'GET':
            pass
            form.name.data = team.name
            form.short_name.data = team.short_name
            form.nickname.data = team.nickname
        if form.validate_on_submit():
            team.name = form.name.data
            team.short_name = form.short_name.data
            if form.nickname.data:
                team.nickname = form.nickname.data
            team.updated_at = now
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                _log_exception(e)
                flash('Error updating team.', 'error')
                return redirect(url_for('main.admin_dashboard'))
            flash('Team Updated.', 'success')
            return redirect(url_for('main.admin_dashboard'))
        return render_template('admin/select_form.html', heading='Create/Edit Team', title='Create team', form=form)
    except Exception as e:
        _log_exception(e)
        flash('Error editing team.', 'error')
        return redirect(url_for('main.admin_dashboard'))


@bp.route('/admin/test_route')
@login_required
def test_route():
    try:
        current_app.logger.info("=== TEST ROUTE CALLED ===")
        return "Test route works!"
    except Exception as e:
        _log_exception(e)
        return "Test route failed", 500



@bp.route('/admin/create_match_week', methods=['GET', 'POST'])
@login_required
def create_match_week():
    
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('main.index'))
    
    # Test database connection
    try:
        weeks = Week.query.order_by(Week.id).all()
        seasons = Season.query.order_by(Season.season_start_year.asc()).all()
        print(f"Found {len(seasons)} seasons in database")
    except Exception as e:
        flash(f'Database error: {str(e)}', 'error')
        return redirect(url_for('main.admin_dashboard'))
    
    week_choices = [(week.id, f"Week {week.week_number}") for week in weeks]
    season_choices = [(season.id, f"{season.season_start_year}-{season.season_end_year}") for season in seasons]
    
    # Get teams for fixture forms
    teams = Team.query.order_by(Team.name).all()
    team_choices = [(team.id, team.name) for team in teams]

    # Initialize form
    if request.method == 'POST':
        # Count the number of fixtures in the POST data
        num_fixtures = int(request.form.get('num_fixtures', 0))
        print(f"Number of fixtures detected in POST data: {num_fixtures}")
        
        # Create a new form class with the correct number of entries
        if num_fixtures > 0:
            # Create a dynamic form class with the right number of entries
            class DynamicCreateMatchWeekForm(CreateMatchWeekForm):
                fixtures = FieldList(FormField(FixtureForm), min_entries=num_fixtures, max_entries=20)
            
            form = DynamicCreateMatchWeekForm()
        else:
            form = CreateMatchWeekForm()
    else:
        form = CreateMatchWeekForm()
    
    # Set choices for the form fields
    form.season.choices = season_choices
    form.week_number.choices = week_choices
    
    # Set team choices for all fixture forms
    for fixture_form in form.fixtures:
        fixture_form.home_team_id.choices = team_choices
        fixture_form.away_team_id.choices = team_choices
        
    # Process form data if it's a POST request
    if request.method == 'POST':
        print(f"Processing form with {len(form.fixtures)} fixtures")
        form.process(request.form)
        print(f"After processing, form has {len(form.fixtures)} fixtures")
        
        # Re-set choices after processing (in case the form was recreated)
        for fixture_form in form.fixtures:
            fixture_form.home_team_id.choices = team_choices
            fixture_form.away_team_id.choices = team_choices

        # Check if form is submitted and valid
    if request.method == 'POST':

        if not form.validate():
            print(f"Form errors: {form.errors}")
            for field_name, errors in form.errors.items():
                print(f"Field {field_name} errors: {errors}")
            print(f"Form data: {form.data}")
        else:
            # Form is valid, process the submission
            for i, fixture_form in enumerate(form.fixtures):
                print(f'Processing fixture {i}: home_team_id={fixture_form.home_team_id.data}, away_team_id={fixture_form.away_team_id.data}')
                if fixture_form.home_team_id.data and fixture_form.away_team_id.data:
                    home_team = Team.query.get(fixture_form.home_team_id.data)
                    away_team = Team.query.get(fixture_form.away_team_id.data)
                    print(f'Fixture {i+1}: {home_team.name if home_team else "Unknown"} vs {away_team.name if away_team else "Unknown"}')
            
            if MatchWeek.query.filter_by(week_id=form.week_number.data, season_id=form.season.data).first():
                flash('Match Week and fixtures exist!', 'info')
                return redirect(request.referrer)
            try:
                match_week = MatchWeek(
                    week_id=form.week_number.data,
                    season_id=form.season.data,
                    predictions_open_time=form.predictions_open_time.data,
                    predictions_close_time=form.predictions_close_time.data,
                    
                )
                
                print(f'Creating MatchWeek: week_id={form.week_number.data}, season_id={form.season.data}')
                db.session.add(match_week)
                db.session.flush()  # This gets us the match_week.id
                print(f'MatchWeek created with ID: {match_week.id}')
                
                fixture_count = 0
                for fixture_form in form.fixtures:
                    if fixture_form.home_team_id.data and fixture_form.away_team_id.data:
                        fixture = Fixture(
                            match_week_id=match_week.id,
                            home_team_id=fixture_form.home_team_id.data,
                            away_team_id=fixture_form.away_team_id.data,
                        )
                        db.session.add(fixture)
                        fixture_count += 1
                        home_team = Team.query.get(fixture_form.home_team_id.data)
                        away_team = Team.query.get(fixture_form.away_team_id.data)
                        print(f'Added fixture: {home_team.name if home_team else "Unknown"} vs {away_team.name if away_team else "Unknown"}')
                
                db.session.commit()
                print(f'Successfully committed {fixture_count} fixtures to database')
                flash(f'Match Week created successfully with {fixture_count} fixtures!', 'success')
                return redirect(url_for('main.admin_dashboard'))
                
            except Exception as e:
                db.session.rollback()
                flash(f'Error creating match week: {str(e)}', 'error')
                return redirect(url_for('main.admin_dashboard'))
    else:
        print("=== FORM VALIDATION FAILED ===")
    return render_template('admin/create_match_week.html', form=form)




@bp.route('/admin/edit_match_week/<int:match_week_id>', methods=['GET', 'POST'])
def edit_match_week(match_week_id):
    try:
        if not current_user.is_admin:
            flash('Access denied!', 'danger')
            return redirect(request.referrer)
        match_week = MatchWeek.query.get_or_404(match_week_id)
        weeks = Week.query.order_by(Week.id).all()
        season = Season.query.filter_by(id=match_week.season_id).first()

        form = MatchWeekUpdateForm()
        form.season.choices = [(season.id, f"{season.season_start_year}/{season.season_end_year}")]
        form.week_number.choices = [(week.id, f"Week {week.week_number}") for week in weeks]
        # form.predictions_open_time.data = match_week.predictions_open_time
        # form.predictions_close_time.data = match_week.predictions_close_time

        if request.method == 'GET':
            form.predictions_open_time.data = match_week.predictions_open_time
            form.predictions_close_time.data = match_week.predictions_close_time
            form.season.data = match_week.season_id
            form.week_number.data = match_week.week_id

        if form.validate_on_submit():
            match_week.season_id = form.season.data
            match_week.week_id = form.week_number.data
            match_week.predictions_open_time = form.predictions_open_time.data
            match_week.predictions_close_time = form.predictions_close_time.data
            # write data to model
            
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                _log_exception(e)
                flash('Error updating match week.', 'error')
                return redirect(url_for('main.admin_dashboard'))
            flash('Match week updated.', 'success')
            return redirect(url_for('main.admin_dashboard'))
        return render_template('admin/form.html', heading='Edit Match Week', title='Edit Match Week', form=form)
    except Exception as e:
        _log_exception(e)
        flash('Error editing match week.', 'error')
        return redirect(url_for('main.admin_dashboard'))



@bp.route('/admin/match_week/delete/<int:match_week_id>', methods=['GET', 'POST'])
def delete_match_week(match_week_id):
    if not current_user.is_admin:
            flash('Access denied!', 'danger')
            return redirect(request.referrer)
    match_week = MatchWeek.query.get_or_404(match_week_id)
    fixture = Fixture.query.get_or_404(match_week_id)
    try:
        db.session.delete(match_week)
        db.session.delete(fixture)
        db.session.commit()
        flash('Match week deleted successfully.', 'success')
        return redirect(url_for('main.admin_dashboard'))
    except Exception as e:
        db.session.rollback()
        _log_exception(e)
        flash(f'Error deleting match week: {str(e)}', 'error')
        return redirect(url_for('main.admin_dashboard'))




@bp.route('/admin/create_season', methods=['GET','POST'])
@login_required
def create_season():
    try:
        if not current_user.is_admin:
            return jsonify({'error': 'Access denied'}), 403
        form = CreateSeasonForm()

        if form.validate_on_submit():
            start = form.start_year.data
            end = form.end_year.data

            if Season.query.filter_by(season_start_year=start, season_end_year=end).first():
                flash('Season exists', 'warning')
                return redirect(request.referrer)
            else:
                try:
                    db.session.add(
                        Season(
                            season_start_year=start,
                            season_end_year=end
                        )
                    )
                    db.session.commit()
                    return redirect(url_for('main.index'))
                except Exception as e:
                    _log_exception(e)
                    #error = jsonify({'error': str(e)}), 500
                    flash(f'Error: {e}', 'danger')
        return render_template('admin/form.html', form=form, heading='Create Season', title='Create Season')
    except Exception as e:
        _log_exception(e)
        flash('Error creating season.', 'error')
        return redirect(url_for('main.admin_dashboard'))



@bp.route('/print-fixtures/<int:match_week_id>')
def print_fixtures(match_week_id):
    try:
        match_week = MatchWeek.query.get_or_404(match_week_id)
        for fx in match_week.fixtures:
            try:
                print(f'{fx.home_team.name} vs {fx.away_team.name}')
            except Exception:
                current_app.logger.exception("Error printing a fixture")
        return render_template('print_fixtures.html', week=match_week)
    except Exception as e:
        _log_exception(e)
        flash('Error printing fixtures.', 'error')
        return redirect(url_for('main.index'))



@bp.route('/admin/fixture/update-single/<int:fixture_id>', methods=['GET','POST'])
@login_required
def update_single_fixture(fixture_id):
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('main.index'))

    fixture = Fixture.query.get_or_404(fixture_id)
    if not fixture:
        flash('Fixture does not exist', 'info')
        return redirect(request.referrer)
    
    form = SingleFixtureForm()

    if request.method == 'GET':
        form.home_team.data = fixture.home_team_id
        form.away_team.data = fixture.away_team_id

    if form.validate_on_submit():
        try:
            # update fixture
            fixture.home_team_id = form.home_team.data
            fixture.away_team_id = form.away_team.data

            # update predictions tied to this fixture
            predictions = Prediction.query.filter_by(fixture_id=fixture.id).all()
            if predictions:
                for prediction in predictions:
                    prediction.home_team_id = fixture.home_team_id
                    prediction.away_team_id = fixture.away_team_id
                    # don't touch home_score_prediction or away_score_prediction
                    # since those are user inputs

            db.session.commit()
            flash('Fixture and related predictions updated successfully.', 'success')
        except Exception as e:
            _log_exception(e)
            db.session.rollback()
            flash(f'An error occurred: {e}', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    return render_template('admin/form.html', form=form, title='Update Fixture', heading='Update Fixture')



@bp.route('/admin/fixture/update-single-score/<int:fixture_id>', methods=['GET','POST'])
@login_required
def update_single_fixture_score(fixture_id):
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('main.index'))
    fixture = Fixture.query.get_or_404(fixture_id)
    if not fixture:
        flash('Fixture does not exist', 'info')
        return redirect(request.referrer)
    
    form = SingleFixtureScoreForm()

    if request.method == 'GET':
        form.home_team.data  = fixture.home_team.name
        form.away_team.data  = fixture.away_team.name
        form.home_score.data    = fixture.home_score
        form.away_score.data    = fixture.away_score

    if form.validate_on_submit():
        try:
            # Only update the score fields...
            fixture.home_score = form.home_score.data
            fixture.away_score = form.away_score.data 
            db.session.commit()
            flash('Fixture updated successfully.', 'success')
        except Exception as e:
            _log_exception(e)
            db.session.rollback()
            flash(f'A error occurred: {e}', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    return render_template('admin/form.html', form=form, title='Update Fixture', heading='Update Team Scores')





@bp.route('/admin/activate_match_week/<int:week_id>', methods=['POST'])
@login_required
def activate_match_week(week_id):
    try:
        if not current_user.is_admin:
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('main.index'))
        MatchWeek.query.update({MatchWeek.is_active: False})
        match_week = MatchWeek.query.get_or_404(week_id)
        match_week.is_active = True
        db.session.commit()
        flash(f'Match Week {match_week.week_number} activated!', 'success')
        return redirect(url_for('main.admin_dashboard'))
    except Exception as e:
        _log_exception(e)
        try:
            db.session.rollback()
        except Exception:
            pass
        flash('Error activating match week.', 'error')
        return redirect(url_for('main.admin_dashboard'))






@bp.route('/predict-matches', methods=['GET', 'POST'])
@login_required
def predict():
    try:
        open_fixtures = Fixture.get_open_for_predictions()
        
        if not open_fixtures:
            flash("Predictions not open at the moment", 'info')
            return redirect(url_for('main.index'))

        # Get match_week_id from the first fixture
        match_week_id = open_fixtures[0].match_week_id
        match_week = MatchWeek.query.filter_by(id=match_week_id).first()
        fixture_id = None
        try:
            fixture_id = Fixture.query.filter_by(match_week_id=match_week_id).first().id
        except Exception:
            fixture_id = None
        
        current_app.logger.info(f"Found {len(open_fixtures)} open fixtures")
        current_app.logger.info(f"Match Week ID: {match_week_id}")
        current_app.logger.info(f"Match Week: {match_week}")

        # Get existing predictions for the current user and match week
        user_predictions = {}
        if current_user.is_authenticated:
            fixture_ids = [fixture.id for fixture in open_fixtures]
            try:
                existing_predictions = Prediction.query.filter(
                    Prediction.user_id == current_user.id,
                    Prediction.fixture_id.in_(fixture_ids)
                ).all()
            except Exception as e:
                existing_predictions = []
                _log_exception(e)
            
            for prediction in existing_predictions:
                user_predictions[prediction.fixture_id] = prediction

        # Create form and populate with correct number of entries
        form = DynamicMatchesForm()
        
        # Populate the form with the correct number of match entries
        while len(form.matches) < len(open_fixtures):
            form.matches.append_entry()
        
        # Remove extra entries if needed
        while len(form.matches) > len(open_fixtures):
            form.matches.pop_entry()
        
        # Populate each form entry with fixture data
        for i, fixture in enumerate(open_fixtures):
            if i < len(form.matches):
                # Pre-populate the team names
                try:
                    form.matches[i].home_team.data = fixture.home_team.name
                except Exception:
                    form.matches[i].home_team.data = ""
                try:
                    form.matches[i].away_team.data = fixture.away_team.name
                except Exception:
                    form.matches[i].away_team.data = ""
                
                # If this is a GET request, pre-populate with existing predictions
                if request.method == 'GET':
                    # Check if user has existing predictions for this fixture
                    if current_user.is_authenticated and fixture.id in user_predictions:
                        prediction = user_predictions[fixture.id]
                        form.matches[i].home_score.data = prediction.home_score_prediction
                        form.matches[i].away_score.data = prediction.away_score_prediction
                    else:
                        # Set default scores to 0 for display
                        form.matches[i].home_score.data = 0
                        form.matches[i].away_score.data = 0

        if request.method == 'POST':
            # Debug: Print form data to see what's being submitted
            current_app.logger.info("Form data received:")
            current_app.logger.info(request.form)
            
            if form.validate_on_submit():
                user_id = current_user.id
                current_app.logger.info(f"User ID: {user_id}")
                
                # Process the submitted data and save to database
                predictions_saved = 0
                
                for i, match_form in enumerate(form.matches):
                    try:
                        # Get the corresponding fixture
                        fixture = open_fixtures[i]
                        fixture_id = fixture.id
                        
                        # Get team IDs
                        home_team = Team.query.filter_by(name=match_form.home_team.data).first()
                        away_team = Team.query.filter_by(name=match_form.away_team.data).first()
                        
                        if not home_team or not away_team:
                            flash(f'Error: Could not find team data for match {i+1}', 'error')
                            continue
                        
                        # Check if prediction already exists for this user and fixture
                        existing_prediction = Prediction.query.filter_by(
                            user_id=user_id, 
                            fixture_id=fixture_id
                        ).first()
                        
                        if existing_prediction:
                            # Update existing prediction
                            existing_prediction.home_team_id = home_team.id
                            existing_prediction.away_team_id = away_team.id
                            existing_prediction.home_score_prediction = match_form.home_score.data
                            existing_prediction.away_score_prediction = match_form.away_score.data
                            existing_prediction.updated_at = datetime.utcnow()
                            current_app.logger.info(f"Updated prediction for fixture {fixture_id}: {home_team.name} {match_form.home_score.data} - {match_form.away_score.data} {away_team.name}")
                        else:
                            # Create new prediction
                            prediction = Prediction(
                                user_id=user_id,
                                fixture_id=fixture_id,
                                home_team_id=home_team.id,
                                away_team_id=away_team.id,
                                home_score_prediction=match_form.home_score.data,
                                away_score_prediction=match_form.away_score.data
                            )
                            db.session.add(prediction)
                            current_app.logger.info(f"Created prediction for fixture {fixture_id}: {home_team.name} {match_form.home_score.data} - {match_form.away_score.data} {away_team.name}")
                        
                        predictions_saved += 1
                    except Exception as e:
                        _log_exception(e)
                        flash('Error processing one of the predictions; continuing.', 'warning')
                
                # Commit all changes to database
                try:
                    db.session.commit()
                    flash(f'{predictions_saved} predictions saved successfully!', 'success')
                except Exception as e:
                    db.session.rollback()
                    _log_exception(e)
                    flash(f'Error saving predictions: {str(e)}', 'error')
                    current_app.logger.error(f"Database error: {e}")

                return redirect(url_for('main.index'))
            else:
                flash('Please correct the errors in the form.', 'error')
        
        return render_template('predict.html', 
                             form=form, 
                             match_week=match_week,
                             user_predictions=user_predictions,
                             fixtures=open_fixtures)
    except Exception as e:
        _log_exception(e)
        flash('An unexpected error occurred while loading predictions.', 'error')
        return redirect(url_for('main.index'))




@bp.route('/predict-matches/get-user/<int:match_week_id>', methods=['GET', 'POST'])
@login_required
def admin_predict_user(match_week_id):

    match_week_id = int(match_week_id)
    if not match_week_id:
        flash('Match week ID is required', 'error')
        return redirect(url_for('main.admin_dashboard'))
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('main.index'))
    
    form = ProxyPredictionsForm()

    if form.validate_on_submit():
        user_id = form.name.data
        print(f"Form submitted with user ID: {user_id}")
        print(f"Selected user ID: {user_id}")
        user = User.query.get(user_id)
        match_week_id = match_week_id
        if not match_week_id:
            flash('Match week not found', 'error')
            return redirect(url_for('main.admin_dashboard'))
        if not user:
            flash('User not found', 'error')
            return redirect(url_for('main.admin_dashboard'))
        return redirect(url_for('main.admin_predict', user_id=user_id, match_week_id=match_week_id))
    
    return render_template('admin/select_form.html', heading='Select User', title='Select User', form=form)




@bp.route('/predict-matches/<int:match_week_id>/user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def admin_predict(match_week_id, user_id):
    """
    Allows an admin to make predictions for a specific user and match week.
    This route fetches fixtures regardless of prediction closure status.
    """
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('main.index'))
    
    try:
        # Get the user and match week objects
        user = User.query.get(user_id)
        if not user:
            flash("User not found.", "error")
            return redirect(url_for('main.admin_dashboard'))

        match_week = MatchWeek.query.get(match_week_id)
        if not match_week:
            flash("Match week not found.", "error")
            return redirect(url_for('main.admin_dashboard'))

        # Fetch all fixtures for the given match week (no time restriction)
        fixtures = Fixture.query.filter_by(match_week_id=match_week_id).order_by(Fixture.id).all()
        if not fixtures:
            flash('No fixtures found for this match week.', 'info')
            return redirect(request.referrer)

        # Get existing predictions for the specified user and match week
        user_predictions = {}
        existing_predictions = Prediction.query.filter_by(user_id=user_id).filter(
            Prediction.fixture_id.in_([f.id for f in fixtures])
        ).all()
        
        for prediction in existing_predictions:
            user_predictions[prediction.fixture_id] = prediction

        # Create form and populate with correct number of entries
        form = DynamicMatchesForm()
        while len(form.matches) < len(fixtures):
            form.matches.append_entry()
        while len(form.matches) > len(fixtures):
            form.matches.pop_entry()

        # Populate each form entry with fixture data
        for i, fixture in enumerate(fixtures):
            if i < len(form.matches):
                form.matches[i].home_team.data = fixture.home_team.name
                form.matches[i].away_team.data = fixture.away_team.name
                
                # Pre-populate with existing predictions on GET
                if request.method == 'GET':
                    if fixture.id in user_predictions:
                        prediction = user_predictions[fixture.id]
                        form.matches[i].home_score.data = prediction.home_score_prediction
                        form.matches[i].away_score.data = prediction.away_score_prediction
                    else:
                        form.matches[i].home_score.data = 0
                        form.matches[i].away_score.data = 0

        if form.validate_on_submit():
            predictions_saved = 0
            for i, match_form in enumerate(form.matches):
                try:
                    fixture = fixtures[i]
                    home_team = Team.query.filter_by(name=match_form.home_team.data).first()
                    away_team = Team.query.filter_by(name=match_form.away_team.data).first()
                    
                    if not home_team or not away_team:
                        flash(f'Error: Could not find team data for match {i+1}', 'error')
                        continue

                    existing_prediction = Prediction.query.filter_by(
                        user_id=user_id,
                        fixture_id=fixture.id
                    ).first()

                    if existing_prediction:
                        existing_prediction.home_score_prediction = match_form.home_score.data
                        existing_prediction.away_score_prediction = match_form.away_score.data
                        existing_prediction.updated_at = datetime.utcnow()
                    else:
                        prediction = Prediction(
                            user_id=user_id,
                            fixture_id=fixture.id,
                            home_team_id=home_team.id,
                            away_team_id=away_team.id,
                            home_score_prediction=match_form.home_score.data,
                            away_score_prediction=match_form.away_score.data
                        )
                        db.session.add(prediction)
                    
                    predictions_saved += 1
                except Exception as e:
                    _log_exception(e)
                    flash('Error processing one of the predictions; continuing.', 'warning')
            
            try:
                db.session.commit()
                flash(f'Predictions saved successfully for {user.name}!', 'success')
            except Exception as e:
                db.session.rollback()
                _log_exception(e)
                flash(f'Error saving predictions: {str(e)}', 'error')
            
            return redirect(url_for('main.admin_dashboard', user_id=user_id, match_week_id=match_week_id))
        
        return render_template('admin/admin_predict.html', 
                               form=form, 
                               match_week=match_week,
                               user_predictions=user_predictions,
                               fixtures=fixtures,
                               user=user)

    except Exception as e:
        _log_exception(e)
        flash('An unexpected error occurred while loading predictions for admin.', 'error')
        return redirect(url_for('main.admin_dashboard'))





@bp.route('/leaderboard')
@login_required
def leaderboard():
    try:
        from sqlalchemy import desc
        users = User.query.order_by(User.total_points.desc()).all()
        for user in users:
            user_scores = 0
            try:
                user_matchweekpoints = MatchWeekPoint.query.filter_by(user_id=user.id).all()
            except Exception:
                user_matchweekpoints = []
            for matchweekpoint in user_matchweekpoints:
                try:
                    user_scores += matchweekpoint.points
                except Exception:
                    pass
            user.total_points = user_scores
        try:
            db.session.commit()
        except Exception as e:
            current_app.logger.error(f"Error updating user scores: {e}")
            _log_exception(e)
            flash('Error updating leaderboard scores', 'error')
            return redirect(url_for('main.index'))

        return render_template('leaderboard.html', users=users, title='Leaderboard')
    except Exception as e:
        _log_exception(e)
        flash('Error loading leaderboard.', 'error')
        return redirect(url_for('main.index'))



@bp.route('/api/add_fixture_form')
def add_fixture_form():
    try:
        form = FixtureForm()
        teams = Team.query.order_by(Team.name).all()
        team_choices = [(team.id, team.name) for team in teams]
        form.home_team_id.choices = team_choices
        form.away_team_id.choices = team_choices
        return render_template('admin/_fixture_form.html', form=form, index='__INDEX__', team_choices=team_choices)
    except Exception as e:
        _log_exception(e)
        # Return an empty fragment so client doesn't crash
        return render_template('admin/_fixture_form.html', form=FixtureForm(), index='__INDEX__', team_choices=[])


@bp.route('/admin/fixture/create', methods=['GET', 'POST'])
def create_fixture():
    try:
        form = FixtureForm()
        matchform = MatchWeekForm()

        # Populate week and season choices
        week_options = [(week.id, f"Week {week.week_number}") for week in Week.query.order_by(Week.id).all()]
        season_options = [(season.id, f"{season.season_start_year}-{season.season_end_year}") for season in Season.query.order_by(Season.season_start_year.asc()).all()]
        matchform.week_number.choices = week_options
        matchform.season.choices = season_options

        if request.method == 'POST':
            # Fetch data from match week form
            try:
                week_id = request.form.get('week_number')
                season_id = request.form.get('season')
                predictions_open_time = datetime.strptime(request.form.get('predictions_open_time'), '%Y-%m-%dT%H:%M')
                predictions_close_time = datetime.strptime(request.form.get('predictions_close_time'), '%Y-%m-%dT%H:%M')
            except Exception as e:
                _log_exception(e)
                flash('Invalid date/time format provided.', 'error')
                return redirect(request.referrer)

            current_app.logger.info(f"Week ID: {week_id}, Season ID: {season_id}, Open Time: {predictions_open_time}, Close Time: {predictions_close_time}")

            # Validate the match week form
            if not week_id or not season_id or not predictions_open_time or not predictions_close_time:
                flash('All fields are required.', 'error')
                return redirect(request.referrer)
            
            if not MatchWeek.query.filter_by(week_id=week_id, season_id=season_id).first():
                try:
                    match_week = MatchWeek(
                        week_id=week_id,
                        season_id=season_id,
                        predictions_open_time=predictions_open_time,
                        predictions_close_time=predictions_close_time
                    )

                    db.session.add(match_week)
                    db.session.flush()

                except Exception as e:
                    _log_exception(e)
                    flash(f'Error: {e}', 'danger')
                    return redirect(request.referrer)

            # Check if the week and season exist
            week = Week.query.get(week_id)
            season = Season.query.get(season_id)
            if not week or not season:
                flash('Invalid week or season selected.', 'error')
                return redirect(request.referrer)
            
            # Collect all fixture pairs from the POST data
            fixtures = []
            i = 0
            while True:
                home_key = f'home_team-{i}'
                away_key = f'away_team-{i}'
                if home_key in request.form and away_key in request.form:
                    home_team = request.form[home_key]
                    away_team = request.form[away_key]
                    try:
                        if not Fixture.query.filter_by(home_team_id=home_team, away_team_id=away_team, match_week_id=match_week.id).first():
                            current_app.logger.info(f"Adding fixture: Home={home_team}, Away={away_team}")
                            db.session.add(
                                Fixture(
                                match_week_id=match_week.id,
                                home_team_id=home_team,
                                away_team_id=away_team
                            ))
                            db.session.commit()
                    except Exception as e:
                        _log_exception(e)
                        flash('Error adding fixture.', 'warning')
                    i += 1
                else:
                    break
            # Print all fixtures (note: original code builds fixtures list but doesn't append; preserve behavior)
            for idx, (home, away) in enumerate(fixtures, 1):
                current_app.logger.info(f"Row {idx}: Home Team = {home}, Away Team = {away}")
            flash(f'{len(fixtures)} fixtures submitted!', 'success')
            return redirect(request.referrer)
        return render_template('admin/create_fixtures.html', fixture_form=form, matchweek_form=matchform, title='Create Fixtures')
    except Exception as e:
        _log_exception(e)
        flash('Error creating fixtures.', 'error')
        return redirect(url_for('main.admin_dashboard'))


@bp.route('/admin/fixture/select-match-week', methods=['GET', 'POST'])
def select_fixture_matchweek():
    try:
        season = Season.query.order_by(Season.id.desc()).first()
        weeks = Week.query.order_by(Week.id).all()
        form = ViewGameWeekPredictionForm()

        form.season.choices = [(season.id, f"{season.season_start_year}/{season.season_end_year}" )] if season else []
        form.match_week.choices = [(week.id, f"Week {week.week_number}" ) for week in weeks]

        if form.validate_on_submit():
            week_id = form.match_week.data  # integer
            season_id = form.season.data 

            return redirect(url_for('main.update_fixture', week_id=week_id, season_id=season_id))

        return render_template('admin/form.html', form=form, heading='Select Match Week', title='Select Match Week')
    except Exception as e:
        _log_exception(e)
        flash('Error selecting fixture match week.', 'error')
        return redirect(url_for('main.admin_dashboard'))


@bp.route('/admin/fixture/update/<int:week_id>/<int:season_id>', methods=['GET', 'POST'])
def update_fixture(season_id, week_id):
    try:
        # Get the existing match week to update
        match_week = Week.query.get(week_id)
        season = Season.query.get(season_id)
        
        form = FixtureForm()
        matchform = MatchWeekForm()

        # Populate week and season choices
        week_options = [(week.id, f"Week {week.week_number}") for week in Week.query.order_by(Week.id).all()]
        season_options = [(season.id, f"{season.season_start_year}-{season.season_end_year}") for season in Season.query.order_by(Season.season_start_year.asc()).all()]
        matchform.week_number.choices = week_options
        matchform.season.choices = season_options

        # Pre-populate form with existing data for GET requests
        if request.method == 'GET' and match_week is not None:
            try:
                matchform.week_number.data = match_week.week_id
                matchform.season.data = match_week.season_id
                matchform.predictions_open_time.data = match_week.predictions_open_time
                matchform.predictions_close_time.data = match_week.predictions_close_time
            except Exception:
                pass

        if request.method == 'POST':
            # Fetch data from match week form
            try:
                week_id = request.form.get('week_number')
                season_id = request.form.get('season')
                predictions_open_time = datetime.strptime(request.form.get('predictions_open_time'), '%Y-%m-%dT%H:%M')
                predictions_close_time = datetime.strptime(request.form.get('predictions_close_time'), '%Y-%m-%dT%H:%M')
            except Exception as e:
                _log_exception(e)
                flash('Invalid date/time format provided.', 'error')
                return redirect(request.referrer)

            current_app.logger.info(f"Updating Match Week {match_week.week_id if match_week else 'unknown'}: Week ID: {week_id}, Season ID: {season_id}, Open Time: {predictions_open_time}, Close Time: {predictions_close_time}")

            # Validate the match week form
            if not week_id or not season_id or not predictions_open_time or not predictions_close_time:
                flash('All fields are required.', 'error')
                return redirect(request.referrer)

            # Check if the week and season exist
            week = Week.query.get(week_id)
            season = Season.query.get(season_id)
            if not week or not season:
                flash('Invalid week or season selected.', 'error')
                return redirect(request.referrer)

            try:
                # Update the existing match week
                match_week.week_id = week_id
                match_week.season_id = season_id
                match_week.predictions_open_time = predictions_open_time
                match_week.predictions_close_time = predictions_close_time

                # Delete existing fixtures for this match week
                existing_fixtures = Fixture.query.filter_by(match_week_id=match_week.id).all()
                for fixture in existing_fixtures:
                    db.session.delete(fixture)

                # Collect all fixture pairs from the POST data and create new ones
                fixtures_added = 0
                i = 0
                while True:
                    home_key = f'home_team-{i}'
                    away_key = f'away_team-{i}'
                    if home_key in request.form and away_key in request.form:
                        home_team = request.form[home_key]
                        away_team = request.form[away_key]
                        if home_team and away_team:  # Only add if both teams are selected
                            current_app.logger.info(f"Updating fixture: Home={home_team}, Away={away_team}")
                            db.session.add(
                                Fixture(
                                match_week_id=match_week.id,
                                home_team_id=home_team,
                                away_team_id=away_team
                            ))
                            fixtures_added += 1
                        i += 1
                    else:
                        break

                db.session.commit()
                flash(f'Match week updated successfully! {fixtures_added} fixtures updated.', 'success')
                return redirect(url_for('main.admin_dashboard'))

            except Exception as e:
                db.session.rollback()
                _log_exception(e)
                flash(f'Error updating match week: {e}', 'danger')
                return redirect(request.referrer)

        # Get existing fixtures for display
        existing_fixtures = []
        try:
            existing_fixtures = Fixture.query.filter_by(match_week_id=match_week.id).all() if match_week else []
        except Exception:
            existing_fixtures = []
        
        return render_template('admin/create_fixtures.html', 
                             fixture_form=form, 
                             matchweek_form=matchform, 
                             title='Update Fixtures',
                             match_week=match_week,
                             existing_fixtures=existing_fixtures)
    except Exception as e:
        _log_exception(e)
        flash('Error updating fixtures.', 'error')
        return redirect(url_for('main.admin_dashboard'))




@bp.route('/admin/scores/select-match-week', methods=['GET', 'POST'])
def select_scores_matchweek():
    try:
        season = Season.query.order_by(Season.id.desc()).first()
        weeks = Week.query.order_by(Week.id).all()
        form = ViewGameWeekPredictionForm()

        form.season.choices = [(season.id, f"{season.season_start_year}/{season.season_end_year}" )] if season else []
        form.match_week.choices = [(week.id, f"Week {week.week_number}" ) for week in weeks]

        if form.validate_on_submit():
            week_id = form.match_week.data  # integer
            season_id = form.season.data 

            return redirect(url_for('main.update_scores', week_id=week_id, season_id=season_id))

        return render_template('admin/form.html', form=form, heading='Select Match Week', title='Select Match Week')
    except Exception as e:
        _log_exception(e)
        flash('Error selecting match week for scores.', 'error')
        return redirect(url_for('main.admin_dashboard'))



@bp.route('/admin/fixture/update-scores/<int:week_id>/<int:season_id>', methods=['GET', 'POST'])
def update_scores(week_id, season_id):
    try:
        if not current_user.is_admin:
            flash('Access denied!', 'danger')
            return redirect(url_for('main.index'))

        now = datetime.utcnow()
        week = Week.query.get_or_404(week_id)
        season = Season.query.get_or_404(season_id)
        match_week = MatchWeek.query.filter_by(season_id=season.id, week_id=week.id).first_or_404()

        if now < match_week.predictions_close_time:
            flash('Predictions must be closed before updating scores.', 'info')
            return redirect(url_for('main.index'))

        # Retrieve fixtures from the Fixture model
        fixtures = Fixture.query.filter_by(match_week_id=match_week.id).all()
        if not fixtures:
            flash("No fixtures found for this match week.", 'info')
            return redirect(url_for('main.index'))

        # Create dynamic form
        form = DynamicMatchesForm()

        # Adjust form fields to match number of fixtures
        while len(form.matches) < len(fixtures):
            form.matches.append_entry()
        while len(form.matches) > len(fixtures):
            form.matches.pop_entry()

        if request.method == 'GET':
            # Populate form with fixture data
            for i, fixture in enumerate(fixtures):
                try:
                    form.matches[i].home_team.data = fixture.home_team.name
                except Exception:
                    form.matches[i].home_team.data = ""
                try:
                    form.matches[i].away_team.data = fixture.away_team.name
                except Exception:
                    form.matches[i].away_team.data = ""
                form.matches[i].home_score.data = fixture.home_score or 0
                form.matches[i].away_score.data = fixture.away_score or 0

        elif form.validate_on_submit():
            # Update scores in the Fixture model
            updated = 0
            for i, match_form in enumerate(form.matches):
                try:
                    fixture = fixtures[i]
                    fixture.home_score = match_form.home_score.data
                    fixture.away_score = match_form.away_score.data
                    
                    # Mark fixture as completed if both scores are provided
                    if fixture.home_score is not None and fixture.away_score is not None:
                        fixture.is_completed = True
                    
                    updated += 1
                except Exception:
                    _log_exception("Error updating individual fixture from form")
                    continue
            
            try:
                db.session.commit()
                flash(f'Successfully updated {updated} fixtures.', 'success')
                # Optionally, you can recalculate points for users based on the updated scores
                for fixture in fixtures:
                    if fixture.is_completed:
                        try:
                            predictions = Prediction.query.filter_by(fixture_id=fixture.id).all()
                        except Exception:
                            predictions = []
                        for prediction in predictions:
                            try:
                                if prediction.home_score_prediction == fixture.home_score and prediction.away_score_prediction == fixture.away_score:
                                    prediction.points_earned = 5  # Correct score
                                elif (prediction.home_score_prediction > prediction.away_score_prediction and fixture.home_score > fixture.away_score) or \
                                     (prediction.home_score_prediction < prediction.away_score_prediction and fixture.home_score < fixture.away_score) or \
                                     (prediction.home_score_prediction == prediction.away_score_prediction and fixture.home_score == fixture.away_score):
                                    prediction.points_earned = 3
                                
                                    prediction.points_earned = 0
                            except Exception:
                                _log_exception("Error calculating points for a prediction")
                return redirect(url_for('main.admin_dashboard'))
            except Exception as e:
                db.session.rollback()
                _log_exception(e)
                flash(f'Error updating scores: {e}', 'danger')
                return redirect(url_for('main.admin_dashboard'))

        return render_template(
            'update_scores.html',
            form=form,
            title='Update Scores',
            heading='Update Fixture Scores',
            match_week=match_week,
            fixtures=fixtures
        )
    except Exception as e:
        _log_exception(e)
        flash('Error updating scores.', 'error')
        return redirect(url_for('main.admin_dashboard'))







@bp.route('/view-predictions', methods=['GET', 'POST'])
def view_prediction():
    try:
        user_id = getattr(current_user, 'id', None)
        form = ViewGameWeekPredictionForm()
        predictions = []
        fixtures = []
        selected_season = None
        selected_match_week = None

        # Get all seasons for dropdown
        seasons = Season.query.order_by(Season.season_start_year.desc()).all()
        season_choices = [(s.id, f"{s.season_start_year}/{s.season_end_year}") for s in seasons]
        
        # Get the latest season to show match weeks for
        latest_season = Season.query.order_by(Season.season_start_year.desc()).first()
        
        if latest_season:
            # Get match weeks for the latest season
            match_weeks = MatchWeek.query.filter_by(season_id=latest_season.id).order_by(MatchWeek.week_id).all()
            week_options = [(mw.id, f"Week {mw.week.week_number}") for mw in match_weeks]
        else:
            week_options = []


        form.season.choices = season_choices
        form.match_week.choices = week_options
        
        if form.validate_on_submit():
            selected_season_id = form.season.data
            selected_match_week_id = form.match_week.data
            
            # Get the selected season and match week
            selected_season = Season.query.get(selected_season_id)
            selected_match_week = MatchWeek.query.get(selected_match_week_id)
            
            if selected_match_week:
                # Get fixtures for the selected match week
                try:
                    fixtures = Fixture.query.filter_by(match_week_id=selected_match_week_id).all()
                except Exception:
                    fixtures = []
                
                # Get all predictions for these fixtures
                fixture_ids = [fixture.id for fixture in fixtures]
                try:
                    predictions = Prediction.query.filter(Prediction.fixture_id.in_(fixture_ids)).filter_by(user_id=user_id).all()
                except Exception:
                    predictions = []
                
                current_app.logger.info(f"Found {len(fixtures)} fixtures and {len(predictions)} predictions for match week {selected_match_week_id}")
        
        return render_template('view_predictions.html', 
                             form=form, 
                             predictions=predictions, 
                             fixtures=fixtures,
                             selected_season=selected_season,
                             selected_match_week=selected_match_week)
    except Exception as e:
        _log_exception(e)
        flash('Error viewing predictions.', 'error')
        return redirect(url_for('main.index'))
    


@bp.route('/predictions/export/<int:match_week_id>', methods=['GET'])
@login_required
def export_predictions(match_week_id):

    if not current_user.is_admin:
        flash('Access denied!', 'danger')
        return redirect(url_for('main.index'))

    try:
        # 1. Fetch data from the database
        fixtures = Fixture.query.filter_by(match_week_id=match_week_id).all()
        if not fixtures:
            flash('No fixtures found for this match week.', 'info')
            return redirect(request.referrer)

        fixture_ids = [f.id for f in fixtures]
        all_predictions = Prediction.query.filter(Prediction.fixture_id.in_(fixture_ids)).all()
        
        # 2. Prepare the data for transposition and Excel export
        
        # Get a consistent, ordered list of fixture keys for the columns
        fixture_keys = [f'{f.home_team.short_name} vs {f.away_team.short_name}' for f in fixtures]
        
        # Create a dictionary to hold all data, with users and system as keys
        data_to_export = {}
        
        # Add the system's predictions as a special entry
        system_predictions = {
            f'{f.home_team.short_name} vs {f.away_team.short_name}': f'{f.home_score}-{f.away_score}' 
            for f in fixtures
        }
        data_to_export['System'] = system_predictions
        
        # Add each user's predictions
        for prediction in all_predictions:
            user_name = f'{prediction.user.name} | {prediction.user.nickname}' if prediction.user.name else 'N/A'
            fixture_key = f'{prediction.fixture.home_team.short_name} vs {prediction.fixture.away_team.short_name}'
            
            if user_name not in data_to_export:
                data_to_export[user_name] = {}
            
            data_to_export[user_name][fixture_key] = f'{prediction.home_score_prediction}-{prediction.away_score_prediction}'
            
        # 3. Create the Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"Match Week {match_week_id}"

        # 4. Write the header row
        headers = ['User'] + fixture_keys
        worksheet.append(headers)
        
        # 5. Write the data rows
        # The 'System' row goes first
        system_row_data = ['System'] + [data_to_export['System'].get(key, 'N/A') for key in fixture_keys]
        worksheet.append(system_row_data)
        
        # Loop through each user and add their prediction row
        for user_name, user_predictions in data_to_export.items():
            if user_name == 'System':
                continue # Skip the system row as it's already added
            
            user_row_data = [user_name] + [user_predictions.get(key, 'N/A') for key in fixture_keys]
            worksheet.append(user_row_data)

        # 6. Save the Excel file to an in-memory BytesIO object
        excel_output = io.BytesIO()
        workbook.save(excel_output)
        excel_output.seek(0)

        # 7. Return the file for download with the correct mimetype
        return send_file(
            excel_output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'predictions_matchweek_{match_week_id}.xlsx'
        )

    except Exception as e:
        _log_exception(e) # Assuming this function is defined for logging
        flash('Error exporting predictions.', 'error')
        return redirect(request.referrer)




@bp.route('/select-scoring-matchweek', methods=['GET', 'POST'])
def select_scoring_matchweek():
    try:
        form = MatchWeekForm()
        return render_template('admin/form.html', heading='Select Match Week', title='Select Match Week', form=form)
    except Exception as e:
        _log_exception(e)
        flash('Error selecting scoring match week.', 'error')
        return redirect(url_for('main.admin_dashboard'))



@bp.route('/generate-matchweek-points/<int:match_week_id>', methods=['GET', 'POST'])
@login_required
def generate_matchweek_points(match_week_id):
    try:
        if current_user.is_anonymous or not current_user.is_admin:
            flash('Access denied!', 'danger')
            return redirect(url_for('main.index'))
        
        match_week = MatchWeek.query.get_or_404(match_week_id)

        fixtures = Fixture.query.filter_by(match_week_id=match_week_id).all()
        if not fixtures:
            flash('No fixtures found for this match week.', 'info')
            return redirect(request.referrer or url_for('main.admin_dashboard'))

        # ensure results are updated
        if any(fx.home_score is None or fx.away_score is None for fx in fixtures):
            flash('Fixtures not updated with results.', 'info')
            return redirect(request.referrer or url_for('main.admin_dashboard'))

        users = User.query.order_by(User.id).all()
        for user in users:
            try:
                user_score = 0

                # ✅ Only get predictions linked to this match_week
                user_predictions = (
                    Prediction.query
                    .join(Fixture, Fixture.id == Prediction.fixture_id)
                    .filter(Prediction.user_id == user.id, Fixture.match_week_id == match_week_id)
                    .all()
                )

                if not user_predictions:
                    continue

                for prediction in user_predictions:
                    fixture = prediction.fixture  # use relationship instead of re-query

                    if fixture.home_score is None or fixture.away_score is None:
                        continue

                    if (fixture.home_score == prediction.home_score_prediction and
                        fixture.away_score == prediction.away_score_prediction):
                        user_score += 5
                    elif ((fixture.home_score == fixture.away_score and prediction.home_score_prediction == prediction.away_score_prediction) or
                          (fixture.home_score < fixture.away_score and prediction.home_score_prediction < prediction.away_score_prediction) or
                          (fixture.home_score > fixture.away_score and prediction.home_score_prediction > prediction.away_score_prediction)):
                        user_score += 3

                # ✅ Update-or-create instead of skipping
                match_week_point = MatchWeekPoint.query.filter_by(
                    user_id=user.id, match_week_id=match_week_id
                ).first()

                if match_week_point:
                    match_week_point.points = user_score
                else:
                    match_week_point = MatchWeekPoint(
                        user_id=user.id, match_week_id=match_week_id, points=user_score
                    )
                    db.session.add(match_week_point)

            except Exception:
                _log_exception("Error generating points for a user")
                continue

        try:
            db.session.commit()
            flash('Match week points generated successfully!', 'success')
            return redirect(url_for('main.weekly_leaderboard', match_week_id=match_week_id))
        except Exception as e:
            db.session.rollback()
            _log_exception(e)
            flash(f'Error generating match week points: {str(e)}', 'danger')
            return redirect(url_for('main.admin_dashboard'))

    except Exception as e:
        _log_exception(e)
        flash('Unexpected error generating match week points.', 'error')
        return redirect(url_for('main.admin_dashboard'))



"""
@bp.route('/generate-matchweek-points/<int:match_week_id>', methods=['GET', 'POST'])
@login_required
def generate_matchweek_points(match_week_id):
    try:
        if current_user.is_anonymous or not current_user.is_admin:
            flash('Access denied!', 'danger')
            return redirect(url_for('main.index'))
        
        match_week = MatchWeek.query.get_or_404(match_week_id)

        fixtures = Fixture.query.filter_by(match_week_id=match_week_id).all()
        # guard against empty fixtures list (original code used fixtures[0] which would crash)
        try:
            if not fixtures or fixtures[0].home_score is None or fixtures[0].away_score is None:
                flash('Fixtures not updated with results.', 'info')
                return redirect(request.referrer)
        except IndexError:
            flash('No fixtures found for this match week.', 'info')
            return redirect(request.referrer)
        except Exception:
            _log_exception("Error checking fixture scores")
            flash('Fixtures not updated with results.', 'info')
            return redirect(request.referrer)

        if not fixtures:
            flash('No fixtures found', 'info')
            return redirect(url_for('main.admin_dashboard'))
        for fx in fixtures:
            try:
                current_app.logger.info(f"home {fx.home_team.name}: {fx.home_score}")
            except Exception:
                current_app.logger.exception("Error reading fixture team/score")

        # Generate a list of users...
        users = User.query.order_by(User.id).all()
        for user in users:
            try:
                user_id = user.id
                current_app.logger.info(f"{user.email}")
                user_score = 0
                user_predictions = Prediction.query.filter_by(user_id=user_id).order_by(Prediction.id).all()
                
                # Check if user has any predictions for this match week. If not, continue to next user
                if not user_predictions:
                    continue
                
                for prediction in user_predictions:
                    try:
                        fixture = Fixture.query.filter_by(home_team_id=prediction.home_team_id, away_team_id=prediction.away_team_id).first()

                        if fixture and fixture.home_score == prediction.home_score_prediction and fixture.away_score == prediction.away_score_prediction:
                            user_score += 5
                        elif fixture and ((fixture.home_score == fixture.away_score and prediction.home_score_prediction == prediction.away_score_prediction) or (fixture.home_score < fixture.away_score and 
                            prediction.home_score_prediction < prediction.away_score_prediction) or (fixture.home_score > fixture.away_score and prediction.home_score_prediction > prediction.away_score_prediction)):
                            user_score += 3
                        else:
                            user_score += 0
                    except Exception:
                        _log_exception("Error processing a user's prediction")
                #print(f"{user.name} : {user_score}")

                # Here you would save the user_score to the MatchWeekPoint model
                try:
                    if MatchWeekPoint.query.filter_by(user_id=user_id, match_week_id=match_week_id).first():
                        #flash("Game Week Points already exist!", 'danger')
                        continue
                except Exception:
                    _log_exception("Error checking existing MatchWeekPoint")
                    # continue attempt to create

                # Create a new MatchWeekPoint entry        
                match_week_point = MatchWeekPoint(user_id=user_id, match_week_id=match_week_id, points=user_score)
                db.session.add(match_week_point)
            except Exception:
                _log_exception("Error generating points for a user")
                continue
        try:
            db.session.commit()
            flash('Match week points generated successfully!', 'success')
            return redirect(url_for('main.weekly_leaderboard', match_week_id=match_week_id))
        except Exception as e:
            db.session.rollback()
            _log_exception(e)
            flash(f'Error generating match week points: {str(e)}', 'danger')
            return redirect(url_for('main.admin_dashboard'))
    except Exception as e:
        _log_exception(e)
        flash('Unexpected error generating match week points.', 'error')
        return redirect(url_for('main.admin_dashboard'))
"""


@bp.route('/select-weekly-leaderboard-matchweek', methods=['GET', 'POST'])
def select_weekly_leaderboard_matchweek():
    try:
        form = SelectMatchWeekForm()

        # ✅ Always set choices before validation
        seasons = Season.query.order_by(Season.season_start_year.desc()).all()
        form.season.choices = [(season.id, f"{season.season_start_year}/{season.season_end_year}") for season in seasons]

        weeks = Week.query.order_by(Week.id).all()
        form.match_week.choices = [(week.id, f"Week {week.week_number}") for week in weeks]

        if form.validate_on_submit():
            season_id = form.season.data
            match_week_id = form.match_week.data
            match_week = MatchWeek.query.filter_by(season_id=season_id, week_id=match_week_id).first()
            if not match_week:
                flash('Match week not found for the selected season and week.', 'warning')
                return redirect(request.referrer)

            return redirect(url_for('main.weekly_leaderboard', match_week_id=match_week.id))

        return render_template('admin/form.html', heading='Select Match Week', title='Select Match Week', form=form)
    except Exception as e:
        _log_exception(e)
        flash('Error selecting weekly leaderboard match week.', 'error')
        return redirect(url_for('main.admin_dashboard'))



@bp.route('/leaderboard/weekly/<int:match_week_id>', methods=['GET', 'POST'])
def weekly_leaderboard(match_week_id):
    try:
        match_week = MatchWeek.query.get_or_404(match_week_id)
        weekly_scores = MatchWeekPoint.query.filter_by(match_week_id=match_week_id).order_by(MatchWeekPoint.points.desc()).all()
        users = User.query.order_by(User.name).all()
        if not weekly_scores:
            flash('No weekly scores found', 'info')
            return redirect(url_for('main.index'))
        return render_template('weekly_leaderboard.html', weekly_scores=weekly_scores, title='Weekly Leaderboard', match_week_id=match_week_id,
                            users=users, heading='Weekly Leaderboard', subheading=f'Weekly Leaderboard for Match Week {match_week.week.week_number}')
    except Exception as e:
        _log_exception(e)
        flash('Error loading weekly leaderboard.', 'error')
        return redirect(url_for('main.index'))



@bp.route('/download_scores_excel')
@login_required
def download_scores_excel():
    try:
        if current_user.is_anonymous or not current_user.is_admin:
            flash('Access denied!', 'danger')
            return redirect(url_for('main.index'))

        # If your page is scoped by a week, pass ?week_id=... from the button.
        week_id = request.args.get('week_id', type=int)

        # --- Get the same data you use for the HTML table ---
        # Adjust these queries to match your app logic.
        query = MatchWeekPoint.query
        if week_id:
            query = query.filter_by(week_id=week_id)

        # Ensure ordering by points desc so Rank matches the page
        weekly_scores = query.order_by(MatchWeekPoint.points.desc()).all()

        user_ids = [s.user_id for s in weekly_scores]
        users = User.query.filter(User.id.in_(user_ids)).all()
        user_map = {u.id: u for u in users}

        # Build rows
        rows = []
        for idx, score in enumerate(weekly_scores, start=1):
            u = user_map.get(score.user_id)
            rows.append({
                "Rank": idx,
                "Name": getattr(u, "name", "") if u else "",
                "Nickname": getattr(u, "nickname", "") if u else "",
                "Points": score.points
            })

        # Create Excel in-memory
        output = io.BytesIO()
        df = pd.DataFrame(rows, columns=["Rank", "Name", "Nickname", "Points"])
        try:
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Weekly Scores")

                # Optional: autosize columns
                ws = writer.sheets["Weekly Scores"]
                for col_cells in ws.columns:
                    max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
                    ws.column_dimensions[col_cells[0].column_letter].width = max_len + 2
        except Exception as e:
            _log_exception(e)
            flash('Error generating Excel file.', 'error')
            return redirect(request.referrer or url_for('main.admin_dashboard'))

        output.seek(0)
        filename = f"weekly_scores{'_'+str(week_id) if week_id else ''}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        _log_exception(e)
        flash('Error downloading scores Excel.', 'error')
        return redirect(url_for('main.admin_dashboard'))



@bp.route('/download/weekly_leaderboard_pdf/<int:match_week_id>')
@login_required
def download_weekly_leaderboard_pdf(match_week_id):
    try:
        if current_user.is_anonymous or not current_user.is_admin:
            flash('Access denied!', 'danger')
            return redirect(url_for('main.index'))
        
        match_week = MatchWeek.query.get_or_404(match_week_id)
        if not match_week:
            # Returning a simple text response in this case.
            # In a real app, you might render an error template.
            return "Match week not found.", 404

        weekly_scores = MatchWeekPoint.query.filter_by(match_week_id=match_week_id).order_by(MatchWeekPoint.points.desc()).all()
        users = User.query.order_by(User.name).all()
        
        # Check if there are scores to display.
        if not weekly_scores:
            return "No scores found for this match week.", 404
            
        # Create a dictionary for efficient user lookup
        users_dict = {user.id: user for user in users}
        
        # Initialize FPDF object
        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()
        
        # Add a title and subtitle
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(200, 10, 'Weekly Leaderboard', ln=1, align='C')
        pdf.set_font('Arial', '', 14)
        pdf.cell(200, 10, f'Match Week {match_week.week.week_number}', ln=1, align='C')
        pdf.ln(10)

        # Define table headers and widths
        headers = ['Rank', 'Name', 'Nickname', 'Points']
        col_widths = [20, 65, 65, 30]

        # Draw table headers
        pdf.set_font('Arial', 'B', 12)
        pdf.set_fill_color(45, 0, 46) 
        pdf.set_text_color(255, 255, 255)
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, border=1, ln=0 if i < len(headers)-1 else 1, align='C', fill=True)

        # Draw table rows
        pdf.set_font('Arial', '', 12)
        pdf.set_text_color(0, 0, 0)
        fill = False
        for i, score in enumerate(weekly_scores):
            try:
                user_info = users_dict.get(score.user_id)
                if user_info:
                    row_data = [
                        str(i + 1),
                        user_info.name,
                        user_info.nickname,
                        str(score.points)
                    ]
                    pdf.set_fill_color(240, 240, 240) if fill else pdf.set_fill_color(255, 255, 255)
                    for j, data in enumerate(row_data):
                        pdf.cell(col_widths[j], 10, data, border=1, ln=0 if j < len(row_data)-1 else 1, align='C', fill=True)
                    fill = not fill
            except Exception:
                _log_exception("Error writing a row to PDF")
                continue

        # Output the PDF to a BytesIO object
        pdf_output = io.BytesIO()
        try:
            pdf_bytes = pdf.output(dest='S')
            pdf_output.write(pdf_bytes)
            pdf_output.seek(0)
        except Exception as e:
            _log_exception(e)
            flash('Error generating PDF.', 'error')
            return redirect(url_for('main.admin_dashboard'))
        
        # Create a Flask response object
        response = make_response(pdf_output.getvalue())
        
        # Set the headers for the PDF download
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=weekly_leaderboard_match_week_{match_week_id}.pdf'
        
        return response
    except Exception as e:
        _log_exception(e)
        flash('Error downloading weekly leaderboard PDF.', 'error')
        return redirect(url_for('main.admin_dashboard'))


@bp.route('/download/leaderboard_pdf')
@login_required
def download_leaderboard_pdf():
    try:
        if current_user.is_anonymous or not current_user.is_admin:
            flash('Access denied!', 'danger')
            return redirect(url_for('main.index'))
        """
        Generates and downloads a PDF of the weekly leaderboard using FPDF2.
        """
        users = User.query.order_by(User.total_points.desc()).all()

        # Initialize FPDF object
        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()
        
        # Add a title and subtitle
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(200, 10, 'Leaderboard', ln=1, align='C')
        pdf.set_font('Arial', '', 14)
        pdf.cell(200, 10, ln=1, align='C')
        #pdf.cell(200, 10, f'Match Week {match_week.week.week_number}', ln=1, align='C')
        pdf.ln(10)

        # Define table headers and widths
        headers = ['Rank', 'Name', 'Nickname', 'Points']
        col_widths = [20, 65, 65, 30]

        # Draw table headers
        pdf.set_font('Arial', 'B', 12)
        pdf.set_fill_color(45, 0, 46) 
        pdf.set_text_color(255, 255, 255)
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, border=1, ln=0 if i < len(headers)-1 else 1, align='C', fill=True)

        # Draw table rows
        pdf.set_font('Arial', '', 12)
        pdf.set_text_color(0, 0, 0)
        fill = False
        for i, user in enumerate(users):
            try:
                if user:
                    row_data = [
                        str(i + 1),
                        user.name,
                        user.nickname,
                        str(user.total_points)
                    ]
                    pdf.set_fill_color(240, 240, 240) if fill else pdf.set_fill_color(255, 255, 255)
                    for j, data in enumerate(row_data):
                        pdf.cell(col_widths[j], 10, data, border=1, ln=0 if j < len(row_data)-1 else 1, align='C', fill=True)
                    fill = not fill
            except Exception:
                _log_exception("Error writing leaderboard row")
                continue

        # Output the PDF to a BytesIO object
        pdf_output = io.BytesIO()
        try:
            pdf_bytes = pdf.output(dest='S')
            pdf_output.write(pdf_bytes)
            pdf_output.seek(0)
        except Exception as e:
            _log_exception(e)
            flash('Error generating PDF.', 'error')
            return redirect(url_for('main.admin_dashboard'))
        
        # Create a Flask response object
        response = make_response(pdf_output.getvalue())
        
        # Set the headers for the PDF download
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=leaderboard.pdf'
        
        return response
    except Exception as e:
        _log_exception(e)
        flash('Error downloading leaderboard PDF.', 'error')
        return redirect(url_for('main.admin_dashboard'))



##########################
#############################
#### CRUD OPERATIONS ########











